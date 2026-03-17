import os
import csv
import sqlite3
import smtplib
import json
import calendar
import io
import threading
import time as _time
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from flask import (Flask, render_template, request, jsonify,
                   redirect, url_for, flash, send_from_directory, send_file)
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = 'close-checklist-secret-2024'

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
DB_PATH     = os.path.join(BASE_DIR, 'checklist.db')
UPLOADS_DIR = os.path.join(BASE_DIR, 'uploads')
CSV_PATH    = r'c:\Users\mhossam\OneDrive - Journal Technologies\Documents\Close Checklist.csv'

STATUS_OPTIONS = ['Not Started', 'In Progress', 'Complete', 'Overdue']
STATUS_COLORS  = {
    'Not Started': 'secondary',
    'In Progress': 'warning',
    'Complete':    'success',
    'Overdue':     'danger',
}

ALLOWED_EXT = {'pdf','xlsx','xls','csv','docx','doc','txt','png','jpg',
               'jpeg','msg','eml','zip','pptx','ppt','gif'}

UPDATABLE_FIELDS = {'status','due_date','notes','description',
                    'category','frequency','assigned_preparer',
                    'assigned_reviewer','timing'}


# ─── DB helpers ───────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA foreign_keys = ON')
    return conn


def init_db():
    os.makedirs(UPLOADS_DIR, exist_ok=True)
    conn = get_db()
    c    = conn.cursor()

    c.execute('''CREATE TABLE IF NOT EXISTS tasks (
        id                INTEGER PRIMARY KEY AUTOINCREMENT,
        category          TEXT    DEFAULT '',
        operating_cycle   TEXT    DEFAULT '',
        sub_cycle         TEXT    DEFAULT '',
        control_reference TEXT    DEFAULT '',
        description       TEXT    NOT NULL,
        frequency         TEXT    DEFAULT '',
        entity            TEXT    DEFAULT 'Journal Technologies, Inc.',
        assigned_preparer TEXT    DEFAULT '',
        assigned_reviewer TEXT    DEFAULT '',
        timing            TEXT    DEFAULT '',
        due_date          TEXT    DEFAULT '',
        status            TEXT    DEFAULT 'Not Started',
        notes             TEXT    DEFAULT '',
        last_updated      TEXT    DEFAULT ''
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS team_emails (
        name  TEXT PRIMARY KEY,
        email TEXT NOT NULL
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS smtp_config (
        key   TEXT PRIMARY KEY,
        value TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS evidence (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        task_id       INTEGER NOT NULL,
        original_name TEXT    NOT NULL,
        stored_name   TEXT    NOT NULL,
        file_size     INTEGER DEFAULT 0,
        uploaded_at   TEXT,
        FOREIGN KEY (task_id) REFERENCES tasks(id) ON DELETE CASCADE
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS custom_columns (
        id       INTEGER PRIMARY KEY AUTOINCREMENT,
        name     TEXT    NOT NULL UNIQUE,
        col_type TEXT    DEFAULT 'text',
        options  TEXT    DEFAULT '[]',
        position INTEGER DEFAULT 0
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS custom_values (
        task_id   INTEGER NOT NULL,
        column_id INTEGER NOT NULL,
        value     TEXT    DEFAULT '',
        PRIMARY KEY (task_id, column_id)
    )''')

    # ── Period management ──────────────────────────────────────────────────────
    c.execute('''CREATE TABLE IF NOT EXISTS periods (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        period_name TEXT    NOT NULL,
        period_date TEXT    NOT NULL,
        created_at  TEXT,
        is_active   INTEGER DEFAULT 1
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS period_snapshots (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        period_id   INTEGER NOT NULL,
        task_id     INTEGER NOT NULL,
        status      TEXT    DEFAULT 'Not Started',
        due_date    TEXT    DEFAULT '',
        notes       TEXT    DEFAULT '',
        UNIQUE (period_id, task_id),
        FOREIGN KEY (period_id) REFERENCES periods(id) ON DELETE CASCADE
    )''')

    conn.commit()

    # Seed tasks from CSV
    c.execute('SELECT COUNT(*) AS n FROM tasks')
    if c.fetchone()['n'] == 0:
        _import_csv(conn)

    # Seed first period: March 2026
    c.execute('SELECT COUNT(*) FROM periods')
    if c.fetchone()[0] == 0:
        conn.execute(
            "INSERT INTO periods (period_name, period_date, created_at, is_active) VALUES (?,?,?,1)",
            ('March 2026', '2026-03-31', datetime.now().isoformat()),
        )
        conn.commit()

    conn.close()


def _import_csv(conn):
    try:
        with open(CSV_PATH, newline='', encoding='utf-8-sig', errors='replace') as f:
            for row in csv.DictReader(f):
                desc = (row.get('Description') or '').strip()
                if not desc:
                    continue
                conn.execute(
                    '''INSERT INTO tasks
                       (category,operating_cycle,sub_cycle,control_reference,
                        description,frequency,entity,assigned_preparer,
                        assigned_reviewer,timing,status)
                       VALUES (?,?,?,?,?,?,?,?,?,?,'Not Started')''',
                    tuple((row.get(k) or '').strip() for k in (
                        'Category','Operating Cycle','Sub Cycle','Control Reference',
                        'Description','Frequency','Entity',
                        'Assigned Preparer/ Owner','Assigned Reviewer','Timing')),
                )
        conn.commit()
    except FileNotFoundError:
        print(f'WARNING: CSV not found at {CSV_PATH}')


def _update_overdue(conn):
    today = date.today().isoformat()
    conn.execute(
        '''UPDATE tasks SET status='Overdue', last_updated=?
           WHERE due_date!='' AND due_date IS NOT NULL AND due_date<?
             AND status NOT IN ('Complete','Overdue')''',
        (datetime.now().isoformat(), today),
    )
    conn.commit()


def _distinct(c, col):
    c.execute(f"SELECT DISTINCT {col} FROM tasks WHERE {col}!='' ORDER BY {col}")
    return [r[0] for r in c.fetchall()]


def _next_period(period_date_str):
    """Return (name, date_iso) for the month after period_date_str."""
    d = date.fromisoformat(period_date_str)
    ny = d.year + 1 if d.month == 12 else d.year
    nm = 1 if d.month == 12 else d.month + 1
    last = calendar.monthrange(ny, nm)[1]
    nd   = date(ny, nm, last)
    return nd.strftime('%B %Y'), nd.isoformat()


def _get_tasks_for_period(conn, c, sql, params, period_id):
    """Run task query, overlay snapshot data if historical, return task list."""
    c.execute(sql, params)
    tasks = [dict(r) for r in c.fetchall()]

    is_historical = period_id not in ('current', '', None)
    if is_historical and tasks:
        ids = [t['id'] for t in tasks]
        ph  = ','.join('?' * len(ids))
        snaps = c.execute(
            f'SELECT task_id,status,due_date,notes FROM period_snapshots '
            f'WHERE period_id=? AND task_id IN ({ph})',
            [int(period_id)] + ids
        ).fetchall()
        snap_map = {r[0]: dict(r) for r in snaps}
        for t in tasks:
            s = snap_map.get(t['id'])
            if s:
                t['status']   = s['status']
                t['due_date'] = s['due_date'] or ''
                t['notes']    = s['notes'] or ''

    return tasks, is_historical


# ─── Main index ───────────────────────────────────────────────────────────────

@app.route('/')
def index():
    conn = get_db()
    c    = conn.cursor()

    period_id = request.args.get('period_id', 'current')

    # Only run overdue auto-update on current period
    if period_id in ('current', '', None):
        _update_overdue(conn)

    category  = request.args.get('category',  '')
    frequency = request.args.get('frequency', '')
    preparer  = request.args.get('preparer',  '')
    reviewer  = request.args.get('reviewer',  '')
    status    = request.args.get('status',    '')
    timing    = request.args.get('timing',    '')

    sql, params = 'SELECT * FROM tasks WHERE 1=1', []
    if category:  sql += ' AND category=?';          params.append(category)
    if frequency: sql += ' AND frequency=?';         params.append(frequency)
    if preparer:  sql += ' AND assigned_preparer=?'; params.append(preparer)
    if reviewer:  sql += ' AND assigned_reviewer=?'; params.append(reviewer)
    if status:    sql += ' AND status=?';            params.append(status)
    if timing:    sql += ' AND timing=?';            params.append(timing)
    sql += ' ORDER BY category, frequency, id'

    tasks, is_historical = _get_tasks_for_period(conn, c, sql, params, period_id)

    # ── Custom columns ────────────────────────────────────────────────────────
    c.execute('SELECT * FROM custom_columns ORDER BY position, id')
    custom_cols = [dict(r) for r in c.fetchall()]
    for col in custom_cols:
        col['options_list'] = json.loads(col.get('options') or '[]')

    # ── Custom values ─────────────────────────────────────────────────────────
    if tasks:
        ids = [t['id'] for t in tasks]
        ph  = ','.join('?' * len(ids))
        cv_rows = c.execute(
            f'SELECT task_id,column_id,value FROM custom_values WHERE task_id IN ({ph})', ids
        ).fetchall()
        cv_map = {}
        for r in cv_rows:
            cv_map.setdefault(r[0], {})[r[1]] = r[2]
        for t in tasks:
            t['custom'] = cv_map.get(t['id'], {})
    else:
        for t in tasks:
            t['custom'] = {}

    # ── Evidence counts ───────────────────────────────────────────────────────
    if tasks:
        ids = [t['id'] for t in tasks]
        ph  = ','.join('?' * len(ids))
        ec_rows = c.execute(
            f'SELECT task_id,COUNT(*) cnt FROM evidence WHERE task_id IN ({ph}) GROUP BY task_id', ids
        ).fetchall()
        ec_map = {r[0]: r[1] for r in ec_rows}
        for t in tasks:
            t['evidence_count'] = ec_map.get(t['id'], 0)
    else:
        for t in tasks:
            t['evidence_count'] = 0

    # ── Filter options ────────────────────────────────────────────────────────
    categories  = _distinct(c, 'category')
    frequencies = _distinct(c, 'frequency')
    timings     = _distinct(c, 'timing')
    preparers   = _distinct(c, 'assigned_preparer')
    reviewers   = _distinct(c, 'assigned_reviewer')

    c.execute('SELECT status, COUNT(*) cnt FROM tasks GROUP BY status')
    all_stats = {r[0]: r[1] for r in c.fetchall()}

    # For historical view, recalc stats from snapshot
    if is_historical:
        stats = {}
        for t in tasks:
            stats[t['status']] = stats.get(t['status'], 0) + 1
    else:
        stats = all_stats

    # ── Periods ───────────────────────────────────────────────────────────────
    c.execute('SELECT * FROM periods ORDER BY period_date DESC')
    all_periods   = [dict(r) for r in c.fetchall()]
    active_period = next((p for p in all_periods if p['is_active']), None)

    # Preview next period name for the roll-forward modal
    next_period_name = ''
    if active_period:
        next_period_name, _ = _next_period(active_period['period_date'])

    # All unique team members for preparer/reviewer dropdowns
    all_members = sorted(set(preparers) | set(reviewers))

    conn.close()
    return render_template('index.html',
        tasks=tasks,
        categories=categories, frequencies=frequencies,
        preparers=preparers,   reviewers=reviewers, timings=timings,
        all_members=all_members,
        stats=stats,
        custom_cols=custom_cols,
        status_options=STATUS_OPTIONS,
        status_colors=STATUS_COLORS,
        all_periods=all_periods,
        active_period=active_period,
        next_period_name=next_period_name,
        period_id=period_id,
        is_historical=is_historical,
        filters=dict(category=category, frequency=frequency,
                     preparer=preparer, reviewer=reviewer,
                     status=status, timing=timing, period_id=period_id),
    )


# ─── Task CRUD ────────────────────────────────────────────────────────────────

@app.route('/update_task/<int:task_id>', methods=['POST'])
def update_task(task_id):
    data = request.get_json()
    fields, values = [], []
    for key, val in data.items():
        if key in UPDATABLE_FIELDS:
            fields.append(f'{key}=?')
            values.append(val)
    if fields:
        fields.append('last_updated=?')
        values += [datetime.now().isoformat(), task_id]
        conn = get_db()
        conn.execute(f'UPDATE tasks SET {",".join(fields)} WHERE id=?', values)
        conn.commit()
        conn.close()
    return jsonify(success=True)


@app.route('/add_task', methods=['POST'])
def add_task():
    d = request.get_json()
    conn = get_db()
    cur = conn.execute(
        '''INSERT INTO tasks
           (category,description,frequency,assigned_preparer,
            assigned_reviewer,timing,due_date,status)
           VALUES (?,?,?,?,?,?,?,'Not Started')''',
        (d.get('category',''), d.get('description','').strip(),
         d.get('frequency',''), d.get('assigned_preparer',''),
         d.get('assigned_reviewer',''), d.get('timing',''), d.get('due_date','')),
    )
    conn.commit()
    row = dict(conn.execute('SELECT * FROM tasks WHERE id=?', (cur.lastrowid,)).fetchone())
    row['custom'] = {}
    row['evidence_count'] = 0
    conn.close()
    return jsonify(success=True, task=row)


@app.route('/delete_task/<int:task_id>', methods=['POST'])
def delete_task(task_id):
    conn = get_db()
    conn.execute('DELETE FROM tasks WHERE id=?', (task_id,))
    conn.commit()
    conn.close()
    return jsonify(success=True)


@app.route('/bulk_update', methods=['POST'])
def bulk_update():
    d        = request.get_json()
    task_ids = d.get('task_ids', [])
    status   = d.get('status', '')
    due_date = d.get('due_date', '')
    now      = datetime.now().isoformat()
    conn     = get_db()
    for tid in task_ids:
        if status:   conn.execute('UPDATE tasks SET status=?,last_updated=? WHERE id=?',   (status, now, tid))
        if due_date: conn.execute('UPDATE tasks SET due_date=?,last_updated=? WHERE id=?', (due_date, now, tid))
    conn.commit()
    conn.close()
    return jsonify(success=True)


# ─── Period management ────────────────────────────────────────────────────────

@app.route('/roll_forward', methods=['POST'])
def roll_forward():
    conn   = get_db()
    active = conn.execute(
        'SELECT * FROM periods WHERE is_active=1 ORDER BY id DESC LIMIT 1'
    ).fetchone()
    if not active:
        conn.close()
        return jsonify(success=False, error='No active period found.')

    pid  = active['id']
    name = active['period_name']

    # Save snapshot of current task states
    tasks = conn.execute('SELECT id,status,due_date,notes FROM tasks').fetchall()
    for t in tasks:
        conn.execute(
            'INSERT OR REPLACE INTO period_snapshots (period_id,task_id,status,due_date,notes) VALUES (?,?,?,?,?)',
            (pid, t['id'], t['status'], t['due_date'] or '', t['notes'] or ''),
        )

    # Compute and create next period
    next_name, next_iso = _next_period(active['period_date'])
    conn.execute('UPDATE periods SET is_active=0 WHERE id=?', (pid,))
    conn.execute(
        'INSERT INTO periods (period_name,period_date,created_at,is_active) VALUES (?,?,?,1)',
        (next_name, next_iso, datetime.now().isoformat()),
    )

    # Reset all tasks for the new period
    conn.execute(
        "UPDATE tasks SET status='Not Started',due_date='',notes='',last_updated=?",
        (datetime.now().isoformat(),),
    )
    conn.commit()
    conn.close()
    return jsonify(success=True, saved_period=name, next_period=next_name)


@app.route('/period/delete/<int:period_id>', methods=['POST'])
def delete_period(period_id):
    conn = get_db()
    p    = conn.execute('SELECT * FROM periods WHERE id=?', (period_id,)).fetchone()
    if not p:
        conn.close()
        return jsonify(success=False, error='Period not found')
    if p['is_active']:
        conn.close()
        return jsonify(success=False, error='Cannot delete the currently active period')
    conn.execute('DELETE FROM period_snapshots WHERE period_id=?', (period_id,))
    conn.execute('DELETE FROM periods WHERE id=?', (period_id,))
    conn.commit()
    conn.close()
    return jsonify(success=True)


@app.route('/period/reopen/<int:period_id>', methods=['POST'])
def reopen_period(period_id):
    conn   = get_db()
    target = conn.execute('SELECT * FROM periods WHERE id=?', (period_id,)).fetchone()
    if not target:
        conn.close()
        return jsonify(success=False, error='Period not found')
    if target['is_active']:
        conn.close()
        return jsonify(success=False, error='Period is already active')

    active = conn.execute('SELECT * FROM periods WHERE is_active=1 ORDER BY id DESC LIMIT 1').fetchone()

    if active:
        # Save current live state into the active period's snapshot before switching
        tasks = conn.execute('SELECT id,status,due_date,notes FROM tasks').fetchall()
        for t in tasks:
            conn.execute(
                'INSERT OR REPLACE INTO period_snapshots (period_id,task_id,status,due_date,notes) VALUES (?,?,?,?,?)',
                (active['id'], t['id'], t['status'], t['due_date'] or '', t['notes'] or ''),
            )
        conn.execute('UPDATE periods SET is_active=0 WHERE id=?', (active['id'],))

    # Reset all tasks then apply target period's snapshots
    conn.execute("UPDATE tasks SET status='Not Started',due_date='',notes='',last_updated=?",
                 (datetime.now().isoformat(),))
    for snap in conn.execute('SELECT * FROM period_snapshots WHERE period_id=?', (period_id,)).fetchall():
        conn.execute(
            'UPDATE tasks SET status=?,due_date=?,notes=?,last_updated=? WHERE id=?',
            (snap['status'], snap['due_date'] or '', snap['notes'] or '',
             datetime.now().isoformat(), snap['task_id']),
        )

    conn.execute('UPDATE periods SET is_active=1 WHERE id=?', (period_id,))
    conn.commit()
    conn.close()
    return jsonify(success=True, period_name=target['period_name'])


# ─── Evidence ─────────────────────────────────────────────────────────────────

@app.route('/evidence/<int:task_id>')
def list_evidence(task_id):
    conn = get_db()
    rows = conn.execute(
        'SELECT id,original_name,file_size,uploaded_at FROM evidence WHERE task_id=? ORDER BY uploaded_at DESC',
        (task_id,)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/upload_evidence/<int:task_id>', methods=['POST'])
def upload_evidence(task_id):
    files = request.files.getlist('files')
    if not files:
        return jsonify(success=False, error='No files received')
    saved = 0
    conn  = get_db()
    for f in files:
        if not f.filename:
            continue
        ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
        if ext not in ALLOWED_EXT:
            continue
        task_dir = os.path.join(UPLOADS_DIR, str(task_id))
        os.makedirs(task_dir, exist_ok=True)
        stored = f'{datetime.now().strftime("%Y%m%d%H%M%S%f")}_{secure_filename(f.filename)}'
        path   = os.path.join(task_dir, stored)
        f.save(path)
        conn.execute(
            'INSERT INTO evidence (task_id,original_name,stored_name,file_size,uploaded_at) VALUES (?,?,?,?,?)',
            (task_id, f.filename, stored, os.path.getsize(path), datetime.now().isoformat()),
        )
        saved += 1
    conn.commit()
    total = conn.execute('SELECT COUNT(*) FROM evidence WHERE task_id=?', (task_id,)).fetchone()[0]
    conn.close()
    return jsonify(success=True, count=saved, total=total)


@app.route('/download_evidence/<int:ev_id>')
def download_evidence(ev_id):
    conn = get_db()
    row  = conn.execute('SELECT * FROM evidence WHERE id=?', (ev_id,)).fetchone()
    conn.close()
    if not row:
        return 'Not found', 404
    task_dir = os.path.join(UPLOADS_DIR, str(row['task_id']))
    return send_from_directory(task_dir, row['stored_name'],
                               as_attachment=True, download_name=row['original_name'])


@app.route('/delete_evidence/<int:ev_id>', methods=['POST'])
def delete_evidence(ev_id):
    conn = get_db()
    row  = conn.execute('SELECT * FROM evidence WHERE id=?', (ev_id,)).fetchone()
    if row:
        path = os.path.join(UPLOADS_DIR, str(row['task_id']), row['stored_name'])
        if os.path.exists(path):
            os.remove(path)
        conn.execute('DELETE FROM evidence WHERE id=?', (ev_id,))
        conn.commit()
    total = conn.execute('SELECT COUNT(*) FROM evidence WHERE task_id=?', (row['task_id'],)).fetchone()[0]
    conn.close()
    return jsonify(success=True, total=total)


# ─── Custom columns ───────────────────────────────────────────────────────────

@app.route('/custom_columns/add', methods=['POST'])
def add_custom_column():
    d        = request.get_json()
    name     = (d.get('name') or '').strip()
    col_type = d.get('col_type', 'text')
    options  = json.dumps([o.strip() for o in d.get('options', []) if o.strip()])
    if not name:
        return jsonify(success=False, error='Column name is required')
    conn = get_db()
    try:
        conn.execute('INSERT INTO custom_columns (name,col_type,options) VALUES (?,?,?)',
                     (name, col_type, options))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify(success=False, error=f'Column "{name}" already exists')
    conn.close()
    return jsonify(success=True)


@app.route('/custom_columns/rename/<int:col_id>', methods=['POST'])
def rename_custom_column(col_id):
    d    = request.get_json()
    name = (d.get('name') or '').strip()
    if not name:
        return jsonify(success=False, error='Name required')
    conn = get_db()
    conn.execute('UPDATE custom_columns SET name=? WHERE id=?', (name, col_id))
    conn.commit()
    conn.close()
    return jsonify(success=True)


@app.route('/custom_columns/delete/<int:col_id>', methods=['POST'])
def delete_custom_column(col_id):
    conn = get_db()
    conn.execute('DELETE FROM custom_columns WHERE id=?', (col_id,))
    conn.execute('DELETE FROM custom_values WHERE column_id=?', (col_id,))
    conn.commit()
    conn.close()
    return jsonify(success=True)


@app.route('/custom_values/update', methods=['POST'])
def update_custom_value():
    d         = request.get_json()
    task_id   = d.get('task_id')
    column_id = d.get('column_id')
    value     = d.get('value', '')
    conn = get_db()
    conn.execute(
        'INSERT OR REPLACE INTO custom_values (task_id,column_id,value) VALUES (?,?,?)',
        (task_id, column_id, value),
    )
    conn.commit()
    conn.close()
    return jsonify(success=True)


# ─── Export to Excel ──────────────────────────────────────────────────────────

@app.route('/export')
def export():
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        return 'openpyxl not installed. Run: pip install openpyxl', 500

    conn = get_db()
    c    = conn.cursor()

    period_id = request.args.get('period_id', 'current')
    category  = request.args.get('category',  '')
    frequency = request.args.get('frequency', '')
    preparer  = request.args.get('preparer',  '')
    reviewer  = request.args.get('reviewer',  '')
    status    = request.args.get('status',    '')
    timing    = request.args.get('timing',    '')

    sql, params = 'SELECT * FROM tasks WHERE 1=1', []
    if category:  sql += ' AND category=?';          params.append(category)
    if frequency: sql += ' AND frequency=?';         params.append(frequency)
    if preparer:  sql += ' AND assigned_preparer=?'; params.append(preparer)
    if reviewer:  sql += ' AND assigned_reviewer=?'; params.append(reviewer)
    if status:    sql += ' AND status=?';            params.append(status)
    if timing:    sql += ' AND timing=?';            params.append(timing)
    sql += ' ORDER BY category, frequency, id'

    tasks, is_historical = _get_tasks_for_period(conn, c, sql, params, period_id)

    # Custom columns
    c.execute('SELECT * FROM custom_columns ORDER BY position, id')
    custom_cols = [dict(r) for r in c.fetchall()]

    # Custom values
    if tasks:
        ids = [t['id'] for t in tasks]
        ph  = ','.join('?' * len(ids))
        cv_rows = c.execute(f'SELECT task_id,column_id,value FROM custom_values WHERE task_id IN ({ph})', ids).fetchall()
        cv_map  = {}
        for r in cv_rows:
            cv_map.setdefault(r[0], {})[r[1]] = r[2]
        for t in tasks:
            t['custom'] = cv_map.get(t['id'], {})
    else:
        for t in tasks:
            t['custom'] = {}

    # Evidence counts
    if tasks:
        ids    = [t['id'] for t in tasks]
        ph     = ','.join('?' * len(ids))
        ec_rows = c.execute(f'SELECT task_id,COUNT(*) cnt FROM evidence WHERE task_id IN ({ph}) GROUP BY task_id', ids).fetchall()
        ec_map  = {r[0]: r[1] for r in ec_rows}
        for t in tasks:
            t['evidence_count'] = ec_map.get(t['id'], 0)
    else:
        for t in tasks:
            t['evidence_count'] = 0

    # Period label
    period_label = 'Current'
    if is_historical:
        p = c.execute('SELECT period_name FROM periods WHERE id=?', (period_id,)).fetchone()
        if p:
            period_label = p[0]
    else:
        ap = c.execute('SELECT period_name FROM periods WHERE is_active=1').fetchone()
        if ap:
            period_label = ap[0]

    conn.close()

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = period_label[:31]   # sheet name max 31 chars

    # Styles
    HDR_FILL = PatternFill(fill_type='solid', fgColor='1A2035')
    HDR_FONT = Font(color='FFFFFF', bold=True, size=10)
    THIN     = Side(style='thin', color='CCCCCC')
    BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    STATUS_FILLS = {
        'Not Started': PatternFill(fill_type='solid', fgColor='E9ECEF'),
        'In Progress': PatternFill(fill_type='solid', fgColor='FFC107'),
        'Complete':    PatternFill(fill_type='solid', fgColor='198754'),
        'Overdue':     PatternFill(fill_type='solid', fgColor='DC3545'),
    }
    STATUS_FONTS = {
        'Not Started': Font(color='444444', size=10),
        'In Progress': Font(color='212529', bold=True, size=10),
        'Complete':    Font(color='FFFFFF', bold=True, size=10),
        'Overdue':     Font(color='FFFFFF', bold=True, size=10),
    }

    # Period info row
    ws.append([f'Close Checklist — {period_label}   |   Exported: {date.today().strftime("%B %d, %Y")}'])
    ws.cell(1, 1).font = Font(bold=True, size=12, color='1A2035')
    ws.append([])  # blank row

    # Header row (row 3)
    HDR_ROW  = 3
    headers  = ['#', 'Category', 'Description', 'Frequency', 'Preparer',
                'Reviewer', 'Timing', 'Due Date', 'Status', 'Notes', 'Evidence']
    headers += [col['name'] for col in custom_cols]

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=HDR_ROW, column=ci, value=h)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.border    = BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[HDR_ROW].height = 28

    # Data rows
    for ri, t in enumerate(tasks, HDR_ROW + 1):
        row_vals = [
            ri - HDR_ROW,
            t.get('category', ''),
            t.get('description', ''),
            t.get('frequency', ''),
            t.get('assigned_preparer', ''),
            t.get('assigned_reviewer', ''),
            t.get('timing', ''),
            t.get('due_date', ''),
            t.get('status', ''),
            t.get('notes', ''),
            t.get('evidence_count', 0),
        ]
        row_vals += [t['custom'].get(col['id'], '') for col in custom_cols]

        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = BORDER
            cell.alignment = Alignment(vertical='top', wrap_text=(ci == 3))
            cell.font      = Font(size=10)

        # Color status cell
        sv    = t.get('status', 'Not Started')
        sc    = headers.index('Status') + 1
        s_cell = ws.cell(row=ri, column=sc)
        s_cell.fill = STATUS_FILLS.get(sv, PatternFill())
        s_cell.font = STATUS_FONTS.get(sv, Font(size=10))

    # Column widths
    widths = [4, 18, 55, 11, 11, 11, 11, 12, 14, 35, 7]
    widths += [14] * len(custom_cols)
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=HDR_ROW, column=ci).column_letter].width = w

    # Freeze header + left 2 cols
    ws.freeze_panes = 'C4'
    ws.auto_filter.ref = f'A{HDR_ROW}:{ws.cell(HDR_ROW, len(headers)).column_letter}{HDR_ROW}'

    # Stream response
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f'CloseChecklist_{period_label.replace(" ", "_")}_{date.today().isoformat()}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── Email ────────────────────────────────────────────────────────────────────

def _get_smtp(conn):
    return {r[0]: r[1] for r in conn.execute('SELECT key,value FROM smtp_config').fetchall()}

def _get_emails(conn):
    return {r[0]: r[1] for r in conn.execute('SELECT name,email FROM team_emails').fetchall()}

def _send_email(smtp, frm, to, subject, body):
    msg = MIMEMultipart()
    msg['From'], msg['To'], msg['Subject'] = frm, to, subject
    msg.attach(MIMEText(body, 'plain'))
    with smtplib.SMTP(smtp.get('smtp_host',''), int(smtp.get('smtp_port', 587))) as s:
        s.starttls()
        s.login(smtp.get('smtp_user',''), smtp.get('smtp_pass',''))
        s.send_message(msg)


@app.route('/send_reminder/<int:task_id>', methods=['POST'])
def send_reminder(task_id):
    conn     = get_db()
    task     = dict(conn.execute('SELECT * FROM tasks WHERE id=?', (task_id,)).fetchone())
    smtp     = _get_smtp(conn)
    emails   = _get_emails(conn)
    conn.close()
    preparer = task.get('assigned_preparer', '')
    to_email = emails.get(preparer)
    if not to_email:
        return jsonify(success=False, error=f'No email for "{preparer}". Add it in Settings.')
    if not smtp.get('smtp_host') or not smtp.get('smtp_user'):
        return jsonify(success=False, error='SMTP not configured. Go to Settings.')
    subject = f"[Close Checklist] Action needed – {task['category']}"
    body = (f"Hi {preparer},\n\nReminder for the following close checklist item:\n\n"
            f"  Category   : {task['category']}\n"
            f"  Description: {task['description']}\n"
            f"  Frequency  : {task['frequency']}\n"
            f"  Status     : {task['status']}\n"
            f"  Due Date   : {task.get('due_date') or 'not set'}\n\n"
            f"Please log in to update your status.\n\nBest regards,\nClose Checklist System")
    try:
        _send_email(smtp, smtp['smtp_user'], to_email, subject, body)
        return jsonify(success=True, message=f'Reminder sent to {to_email}')
    except Exception as e:
        return jsonify(success=False, error=str(e))


@app.route('/send_reminders_bulk', methods=['POST'])
def send_reminders_bulk():
    d             = request.get_json()
    filter_status = d.get('status', 'Overdue')
    conn          = get_db()
    tasks         = [dict(r) for r in conn.execute('SELECT * FROM tasks WHERE status=?', (filter_status,)).fetchall()]
    smtp          = _get_smtp(conn)
    emails        = _get_emails(conn)
    conn.close()
    if not smtp.get('smtp_host') or not smtp.get('smtp_user'):
        return jsonify(success=False, error='SMTP not configured. Go to Settings.')
    grouped = {}
    for t in tasks:
        grouped.setdefault(t.get('assigned_preparer') or 'Unknown', []).append(t)
    sent, errors = [], []
    for preparer, ptasks in grouped.items():
        to_email = emails.get(preparer)
        if not to_email:
            errors.append(f'No email for {preparer}'); continue
        lines = '\n'.join(f"  • [{t['category']}] {t['description'][:80]}  (Due: {t.get('due_date') or '—'})"
                          for t in ptasks)
        body  = (f"Hi {preparer},\n\nYou have {len(ptasks)} {filter_status.lower()} task(s):\n\n"
                 f"{lines}\n\nPlease log in to update your status.\n\nBest regards,\nClose Checklist System")
        try:
            _send_email(smtp, smtp['smtp_user'], to_email,
                        f"[Close Checklist] {len(ptasks)} {filter_status} task(s) need attention", body)
            sent.append(f'{preparer} ({to_email})')
        except Exception as e:
            errors.append(f'{preparer}: {e}')
    return jsonify(success=True, sent=sent, errors=errors,
                   message=f'Sent to {len(sent)} member(s). {len(errors)} error(s).')


# ─── Settings ─────────────────────────────────────────────────────────────────

@app.route('/settings', methods=['GET', 'POST'])
def settings():
    conn = get_db()

    if request.method == 'POST':
        # Save SMTP
        for key in ('smtp_host', 'smtp_port', 'smtp_user', 'smtp_pass'):
            conn.execute('INSERT OR REPLACE INTO smtp_config (key,value) VALUES (?,?)',
                         (key, request.form.get(key, '').strip()))

        # Save team emails — robust parsing, never silently wipe
        members_json = (request.form.get('team_emails') or '').strip()
        if members_json:
            try:
                members = json.loads(members_json)
                conn.execute('DELETE FROM team_emails')
                for m in members:
                    name  = (m.get('name')  or '').strip()
                    email = (m.get('email') or '').strip()
                    if name and email:
                        conn.execute('INSERT OR REPLACE INTO team_emails (name,email) VALUES (?,?)',
                                     (name, email))
                conn.commit()
            except Exception as e:
                flash(f'Warning: could not save team emails ({e})', 'warning')
        else:
            conn.commit()   # still save SMTP

        conn.close()
        flash('Settings saved!', 'success')
        return redirect(url_for('settings'))

    # GET — build member list from BOTH tasks and existing team_emails
    smtp_config     = _get_smtp(conn)
    existing_emails = _get_emails(conn)

    task_names = {r[0] for r in conn.execute(
        '''SELECT DISTINCT assigned_preparer n FROM tasks WHERE assigned_preparer!=''
           UNION SELECT DISTINCT assigned_reviewer FROM tasks WHERE assigned_reviewer!=''
        ''').fetchall()}

    # Include anyone already saved in team_emails even if not in tasks
    all_names    = sorted(task_names | set(existing_emails.keys()))
    team_members = [{'name': n, 'email': existing_emails.get(n, '')} for n in all_names]

    conn.close()
    return render_template('settings.html', smtp_config=smtp_config, team_members=team_members)


# ─── Scheduled overdue check ──────────────────────────────────────────────────

def _start_scheduler():
    """Background thread: runs _update_overdue() every day at midnight PT."""
    PT = ZoneInfo('America/Los_Angeles')

    def _loop():
        while True:
            now          = datetime.now(PT)
            next_midnight = (now + timedelta(days=1)).replace(
                hour=0, minute=0, second=0, microsecond=0)
            _time.sleep((next_midnight - now).total_seconds())
            try:
                conn = get_db()
                _update_overdue(conn)
                conn.close()
                print(f'[scheduler] Overdue check ran at {datetime.now(PT).strftime("%Y-%m-%d %H:%M PT")}')
            except Exception as e:
                print(f'[scheduler] Error during overdue check: {e}')

    t = threading.Thread(target=_loop, daemon=True)
    t.start()


# ─── Entry point ──────────────────────────────────────────────────────────────

if __name__ == '__main__':
    init_db()
    _start_scheduler()
    print('\n  Close Checklist  →  http://127.0.0.1:5050\n')
    app.run(debug=True, port=5050)
